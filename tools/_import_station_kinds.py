# -*- coding: utf-8 -*-
"""Import station/halt kinds from Участки работы мяэд (станции).xlsx into line-sections.json."""
from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "Участки работы мяэд (станции).xlsx"
OUT = ROOT / "app" / "data" / "line-sections.json"

NAME_MAP = {
    "Москва 3": "Москва-3",
    "Ростокино": "Ростокино (Ярославское направление)",
    "Подлипки Дачные": "Подлипки-Дачные",
    "Зеленый Бор": "Зелёный Бор",
    "Ивантеевка 2": "Ивантеевка-2",
    "Фрязино Товарная": "Фрязино-Тов.",
    "Фрязино Пассажирская": "Фрязино-Пасс.",
    "Александров 1": "Александров-1",
    "Щелково": "Щёлково",
    "Поселок Дальний": "Посёлок Дальний",
    "Федоровское": "Фёдоровское",
    "Лесная": "Лесная (64 км)",
    "Путевой Пост 81 км": "81 км",
    "Ост. Пункт 83 км": "83 км",
    "Шубино": "Шубино (90 км)",
    "Нагорное": "Нагорное (43 км)",
}

TYPE_MAP = {
    "станция": "station",
    "останов. пункт": "halt",
    "остановочный пункт": "halt",
    "оп": "halt",
}


def canon(name: str) -> str:
    name = " ".join(str(name).split())
    return NAME_MAP.get(name, name)


def kind_of(raw) -> str | None:
    if raw is None:
        return None
    key = " ".join(str(raw).split()).casefold().replace("ё", "е")
    return TYPE_MAP.get(key)


def main() -> int:
    if not XLSX.exists():
        print(f"missing {XLSX}", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb.active
    kinds: dict[str, str] = {}

    # name col, type col for each section block
    for name_c, type_c in ((2, 3), (6, 7), (10, 11), (14, 15)):
        for r in range(2, (ws.max_row or 2) + 1):
            name = ws.cell(r, name_c).value
            typ = ws.cell(r, type_c).value
            if not name or not str(name).strip():
                continue
            k = kind_of(typ)
            if not k:
                continue
            title = canon(name)
            # station wins over halt if conflict
            if title in kinds and kinds[title] == "station":
                continue
            kinds[title] = k

    data = json.loads(OUT.read_text(encoding="utf-8"))
    data["station_kinds"] = dict(sorted(kinds.items(), key=lambda item: item[0].casefold()))

    # Ensure every from_moscow name has a kind (default halt if unknown)
    missing = []
    for section in data.get("sections", []):
        for name in section.get("from_moscow", []):
            if name not in data["station_kinds"]:
                missing.append(name)
                data["station_kinds"][name] = "halt"

    OUT.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    stations = sum(1 for v in data["station_kinds"].values() if v == "station")
    halts = sum(1 for v in data["station_kinds"].values() if v == "halt")
    print(f"Wrote station_kinds: {stations} station, {halts} halt, total {len(data['station_kinds'])}")
    if missing:
        print(f"  defaulted missing to halt: {missing}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
