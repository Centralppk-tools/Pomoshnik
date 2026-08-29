#!/usr/bin/env python3
"""Export shift hours from xlsx (Часы смен/) → yandex-cloud/function/data/shift-templates.json"""

from __future__ import annotations

import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
HOURS_DIR = ROOT / "Часы смен"
OUT_PATH = ROOT / "yandex-cloud" / "function" / "data" / "shift-templates.json"
SCHEMA_DIR = ROOT / "tools" / "shift-hours"

DAY_ROUTE_RE = re.compile(r"^[ДD]\d", re.I)
NIGHT_ROUTE_RE = re.compile(r"^[НN]\d", re.I)
MORNING_ROUTE_RE = re.compile(r"^\d+[УU]$", re.I)

FILE_MARKERS = [
    ("ПН_ЧТ", "пн-чт"),
    ("_ПТ", "пт"),
    ("_СБ", "сб"),
    ("_ВС", "вс"),
]

NORMATIVE_PERIODS = [
    {
        "id": "2026-06-24",
        "normativeFrom": "2026-06-24",
        "normativeTo": "2026-07-16",
        "folders": [HOURS_DIR / "с 24_06"],
        "excludeSubdirs": set(),
    },
    {
        "id": "2026-07-17",
        "normativeFrom": "2026-07-17",
        "normativeTo": "2026-08-04",
        "folders": [HOURS_DIR / "c 17_07"],
        "excludeSubdirs": set(),
    },
    {
        "id": "2026-08-05",
        "normativeFrom": "2026-08-05",
        "normativeTo": None,
        "folders": [HOURS_DIR / "c 05_08"],
        "excludeSubdirs": set(),
    },
]

HEADER_ROW = 6
DATA_START_ROW = 7

# Нативный xlsx: ночь с col 11. PDF (pdfplumber): блок сдвинут на −1, обед — одна ячейка.
# Перерыв (обед): у Д — правее «Рабочее время смены» (cols 7–8); у Н — левее «№ марш. утром» (cols 19–20).
LAYOUT_NATIVE = {
    "night_route_col": 11,
    "morning_route_col": 22,
    "day_fields": {
        "startPlace": 2,
        "startTime": 3,
        "trains": 4,
        "endTime": 5,
        "workHours": 6,
        "nightHours": 9,
    },
    "day_lunch": (7, 8),
    "night_fields": {
        "startPlace": 12,
        "startTime": 13,
        "trains": 14,
        "endTime": 15,
        "workHours": 16,
        "nightHours": 17,
        "morningRoute": 21,
    },
    "night_lunch": (19, 20),
    "morning_fields": {
        "startPlace": 23,
        "startTime": 24,
        "trains": 25,
        "endTime": 26,
        "workHours": 27,
        "nightHours": 28,
    },
}

LAYOUT_PDF = {
    "night_route_col": 10,
    "morning_route_col": 20,
    "day_fields": {
        "startPlace": 2,
        "startTime": 3,
        "trains": 4,
        "endTime": 5,
        "workHours": 6,
        "nightHours": 8,
    },
    "day_lunch": (7,),
    "night_fields": {
        "startPlace": 11,
        "startTime": 12,
        "trains": 13,
        "endTime": 14,
        "workHours": 15,
        "nightHours": 16,
        "morningRoute": 19,
    },
    "night_lunch": (18,),
    "morning_fields": {
        "startPlace": 21,
        "startTime": 22,
        "trains": 23,
        "endTime": 24,
        "workHours": 25,
        "nightHours": 26,
    },
}


def cell_str(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return text


def normalize_route(route: str) -> str:
    return cell_str(route).upper().replace(" ", "")


def normalize_lunch(value) -> str:
    text = cell_str(value)
    if not text:
        return ""
    if text in ("*", "**", "***"):
        return text
    last = text.split("\n")[-1].strip()
    return last if last in ("*", "**", "***") else ""


TIME_IN_CELL_RE = re.compile(r"(\d{1,2}:\d{2})")


def extract_times_from_cell(value) -> list[str]:
    text = cell_str(value).replace("\n", " ")
    return TIME_IN_CELL_RE.findall(text)


def extract_lunch_marker(*values) -> str:
    for value in values:
        text = cell_str(value)
        if not text:
            continue
        if "***" in text:
            return "***"
        if "**" in text:
            return "**"
        if "*" in text:
            return "*"
    return ""


def format_break_lunch(*cells) -> str:
    """Перерыв: две ячейки (нативный xlsx) или одна (PDF: «15:00 16:00 **»)."""
    if not cells:
        return ""

    all_times: list[str] = []
    marker = ""
    for cell in cells:
        all_times.extend(extract_times_from_cell(cell))
        marker = marker or extract_lunch_marker(cell) or normalize_lunch(cell)

    if len(all_times) >= 2:
        lunch = f"{all_times[0]}-{all_times[1]}"
        if marker:
            lunch = f"{lunch} {marker}"
        return lunch

    if len(cells) >= 2:
        start_times = extract_times_from_cell(cells[0])
        end_times = extract_times_from_cell(cells[1])
        if start_times and end_times:
            lunch = f"{start_times[0]}-{end_times[0]}"
            if marker:
                lunch = f"{lunch} {marker}"
            return lunch

    if marker:
        return marker
    if len(all_times) == 1:
        return all_times[0]
    return ""


def detect_marker(path: Path) -> str:
    name = path.name.upper()
    for pattern, marker in FILE_MARKERS:
        if pattern.upper() in name:
            return marker
    raise ValueError(f"Cannot detect weekday marker for {path.name}")


def list_xlsx_files(folder: Path, exclude_subdirs: set[str]) -> list[Path]:
    files: list[Path] = []
    for path in sorted(folder.rglob("*.xlsx")):
        rel_parts = path.relative_to(folder).parts
        if rel_parts and rel_parts[0] in exclude_subdirs:
            continue
        if path.name.startswith("~$") or path.name.startswith("_"):
            continue
        files.append(path)
    return files


def read_row_segment(ws, row: int, mapping: dict[str, int]) -> dict[str, str]:
    out: dict[str, str] = {}
    for key, col in mapping.items():
        out[key] = cell_str(ws.cell(row, col).value)
    return out


def flat_row(date: str, route: str, fields: dict[str, str]) -> dict[str, str]:
    row = {
        "date": date,
        "route": normalize_route(route),
    }
    for key in ("startPlace", "startTime", "trains", "endTime", "workHours", "nightHours", "morningRoute", "lunch"):
        value = cell_str(fields.get(key, ""))
        if value:
            row[key] = value
    return row


def find_first_data_row(ws) -> int:
    max_row = ws.max_row or DATA_START_ROW
    for row in range(HEADER_ROW, max_row + 1):
        day_route = cell_str(ws.cell(row, 1).value)
        if DAY_ROUTE_RE.match(day_route):
            return row
    return DATA_START_ROW


ISO_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")
WEEKDAY_MARKERS = {"пн-чт", "пт", "сб", "вс"}


def iso_to_weekday_marker(iso: str) -> str:
    """YYYY-MM-DD → группа дня (пн-чт / пт / сб / вс), как в app/index.html."""
    from datetime import date as date_cls

    y, m, d = map(int, iso.split("-"))
    dow = date_cls(y, m, d).weekday()  # 0=пн … 6=вс
    if dow <= 3:
        return "пн-чт"
    if dow == 4:
        return "пт"
    if dow == 5:
        return "сб"
    return "вс"


def normalize_day_marker(value: str) -> str:
    """Маркер или ISO-дата → пн-чт / пт / сб / вс."""
    marker = cell_str(value)
    if marker in WEEKDAY_MARKERS:
        return marker
    if ISO_DATE_RE.match(marker):
        return iso_to_weekday_marker(marker)
    raise ValueError(f"Неизвестный маркер дня: {marker}")


def detect_workbook_layout(ws, data_start_row: int, max_row: int) -> dict:
    """PDF-таблица сдвигает ночной блок на 1 колонку влево (col 10 вместо 11)."""
    native_hits = 0
    pdf_hits = 0
    for row in range(data_start_row, max_row + 1):
        if not DAY_ROUTE_RE.match(cell_str(ws.cell(row, 1).value)):
            continue
        if NIGHT_ROUTE_RE.match(cell_str(ws.cell(row, LAYOUT_NATIVE["night_route_col"]).value)):
            native_hits += 1
        if NIGHT_ROUTE_RE.match(cell_str(ws.cell(row, LAYOUT_PDF["night_route_col"]).value)):
            pdf_hits += 1
    if pdf_hits > native_hits:
        return LAYOUT_PDF
    return LAYOUT_NATIVE


def parse_workbook_with_marker(path: Path, marker: str) -> list[dict[str, str]]:
    marker = normalize_day_marker(marker)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active

    morning_index: dict[str, int] = {}
    night_rows: list[tuple[int, str]] = []
    flat_rows: list[dict[str, str]] = []

    data_start_row = find_first_data_row(ws)
    max_row = ws.max_row or data_start_row
    layout = detect_workbook_layout(ws, data_start_row, max_row)
    morning_col = layout["morning_route_col"]

    for row in range(data_start_row, max_row + 1):
        day_route = cell_str(ws.cell(row, 1).value)
        if not DAY_ROUTE_RE.match(day_route):
            continue

        morning_route_col = cell_str(ws.cell(row, morning_col).value)
        if MORNING_ROUTE_RE.match(morning_route_col):
            morning_index[normalize_route(morning_route_col)] = row

        day_fields = read_row_segment(ws, row, layout["day_fields"])
        day_lunch_cols = layout["day_lunch"]
        day_fields["lunch"] = format_break_lunch(
            *[ws.cell(row, col).value for col in day_lunch_cols]
        )
        flat_rows.append(flat_row(marker, day_route, day_fields))

        night_route = cell_str(ws.cell(row, layout["night_route_col"]).value)
        if not NIGHT_ROUTE_RE.match(night_route):
            continue

        night_fields = read_row_segment(ws, row, layout["night_fields"])
        night_lunch_cols = layout["night_lunch"]
        night_fields["lunch"] = format_break_lunch(
            *[ws.cell(row, col).value for col in night_lunch_cols]
        )
        flat_rows.append(flat_row(marker, night_route, night_fields))

        morning_route = cell_str(night_fields.get("morningRoute", ""))
        if morning_route:
            night_rows.append((row, normalize_route(morning_route)))

    for _source_row, morning_route in night_rows:
        target_row = morning_index.get(morning_route)
        if not target_row:
            continue

        morning_fields = read_row_segment(ws, target_row, layout["morning_fields"])
        flat_rows.append(flat_row(marker, morning_route, morning_fields))

    wb.close()
    return dedupe_rows(flat_rows)


def parse_workbook(path: Path) -> list[dict[str, str]]:
    return parse_workbook_with_marker(path, detect_marker(path))


def route_kind(route: str) -> str:
    r = normalize_route(route)
    if re.match(r"^[ДD]", r):
        return "day"
    if re.match(r"^[НN]", r):
        return "night"
    if MORNING_ROUTE_RE.match(r):
        return "morn"
    return "other"


def analyze_shift_rows(rows: list[dict[str, str]]) -> list[dict]:
    """Предупреждения для превью Naryad_pan (совместимо с findShiftTemplate)."""
    warnings: list[dict] = []
    by_key = {(r["date"], r["route"]): r for r in rows}
    routes_by_date: dict[str, set[str]] = {}
    for row in rows:
        routes_by_date.setdefault(row["date"], set()).add(row["route"])

    seen: set[tuple[str, str]] = set()
    for row in rows:
        key = (row["date"], row["route"])
        if key in seen:
            warnings.append({
                "level": "error",
                "code": "DUPLICATE_ROW",
                "date": row["date"],
                "route": row["route"],
                "message": f"Дубликат {row['route']} на «{row['date']}»",
            })
        seen.add(key)

        if not cell_str(row.get("trains")):
            warnings.append({
                "level": "warn",
                "code": "EMPTY_TRAINS",
                "date": row["date"],
                "route": row["route"],
                "message": f"{row['route']}: нет поездов",
            })

        mr = cell_str(row.get("morningRoute"))
        if mr and route_kind(row["route"]) == "night":
            if (row["date"], normalize_route(mr)) not in by_key:
                warnings.append({
                    "level": "error",
                    "code": "MISSING_MORNING",
                    "date": row["date"],
                    "route": row["route"],
                    "message": f"Н{row['route'][1:]} → {mr}, но строки {mr} нет",
                })

    for row in rows:
        if route_kind(row["route"]) == "morn" and not cell_str(row.get("trains")):
            continue
        if route_kind(row["route"]) == "day" and not cell_str(row.get("workHours")):
            warnings.append({
                "level": "warn",
                "code": "EMPTY_WORK_HOURS",
                "date": row["date"],
                "route": row["route"],
                "message": f"{row['route']}: пустые рабочие часы",
            })

    return warnings


def row_with_kind(row: dict[str, str]) -> dict[str, str]:
    out = dict(row)
    out["kind"] = route_kind(row.get("route", ""))
    return out


def dedupe_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    seen: dict[tuple[str, str], dict[str, str]] = {}
    for row in rows:
        key = (row["date"], row["route"])
        seen[key] = row
    return list(seen.values())


def sort_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    marker_order = {"пн-чт": 0, "пт": 1, "сб": 2, "вс": 3}

    def route_sort(route: str) -> tuple:
        kind = 0
        if route.startswith("Д"):
            kind = 0
        elif route.startswith("Н"):
            kind = 1
        else:
            kind = 2
        num = re.sub(r"[^\d]", "", route)
        return (kind, int(num) if num.isdigit() else 999, route)

    return sorted(
        rows,
        key=lambda row: (
            marker_order.get(row["date"], 99),
            route_sort(row["route"]),
        ),
    )


def export_period(period: dict) -> dict:
    rows: list[dict[str, str]] = []
    source_files: list[str] = []

    for folder in period["folders"]:
        if not folder.exists():
            print(f"warn: folder missing: {folder}", file=sys.stderr)
            continue
        for path in list_xlsx_files(folder, period["excludeSubdirs"]):
            parsed = parse_workbook(path)
            rows.extend(parsed)
            source_files.append(str(path.relative_to(ROOT)).replace("\\", "/"))
            print(f"  {path.name}: {len(parsed)} rows")

    merged = sort_rows(dedupe_rows(rows))
    routes = sorted({row["route"] for row in merged})

    return {
        "id": period["id"],
        "normativeFrom": period["normativeFrom"],
        "normativeTo": period["normativeTo"],
        "sourceFiles": source_files,
        "shiftDetails": merged,
        "stats": {
            "rows": len(merged),
            "routes": len(routes),
            "markers": sorted({row["date"] for row in merged}),
        },
    }


def main() -> int:
    print("Exporting shift hours…")
    normatives = []
    all_rows: list[dict[str, str]] = []

    for period in NORMATIVE_PERIODS:
        print(f"\n[{period['id']}]")
        bundle = export_period(period)
        normatives.append(bundle)
        all_rows.extend(bundle["shiftDetails"])

    payload = {
        "version": normatives[-1]["id"] if normatives else "0.0.0",
        "generatedAt": datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        "source": {
            "type": "local-xlsx",
            "folder": "Часы смен/",
            "exporter": "tools/export-shifts.py",
        },
        "normatives": [
            {
                "id": item["id"],
                "normativeFrom": item["normativeFrom"],
                "normativeTo": item["normativeTo"],
                "sourceFiles": item["sourceFiles"],
                "stats": item["stats"],
                "shiftDetails": item["shiftDetails"],
            }
            for item in normatives
        ],
        "shiftDetails": sort_rows(dedupe_rows(all_rows)),
        "stats": {
            "rows": len(all_rows),
            "routes": len({row["route"] for row in all_rows}),
            "markers": sorted({row["date"] for row in all_rows}),
            "normatives": len(normatives),
        },
    }

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"\nWrote {OUT_PATH} ({payload['stats']['rows']} rows, {payload['stats']['routes']} routes)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
